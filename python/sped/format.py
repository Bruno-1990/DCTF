# format.py — toda a formatação do XLSX centralizada (versão unificada)

from __future__ import annotations
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configurações visuais
HEADER_FONT   = Font(bold=True)
HEADER_HEIGHT = 28
ROW_HEIGHT    = 22

# Cores (ARGB)
HEADER_FILL = PatternFill(start_color="FFC5D9F1", end_color="FFC5D9F1", fill_type="solid")  # C5D9F1
ZEBRA_FILL  = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")  # DCE6F1

# -------------------------
# Funções auxiliares
# -------------------------

def _auto_width(ws: Worksheet) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[letter].width = min(max(8, max_len + 2), 60)

def _center_all(ws: Worksheet) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = cell.value.replace("\\n", " ").replace("\\r", " ")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

def _header_and_row_heights(ws: Worksheet) -> None:
    if ws.max_row >= 1:
        ws.row_dimensions[1].height = HEADER_HEIGHT
        for c in ws[1]:
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
    for r in range(2, (ws.max_row or 1) + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT

def _freeze_header(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"  # congela linha 1

def _highlight_delta_headers(ws: Worksheet) -> None:
    """Destaca cabeçalhos que começam com 'Delta' (case-sensitive)."""
    fill = PatternFill(start_color="FFFDE9D9", end_color="FFFDE9D9", fill_type="solid")
    if ws.max_row >= 1:
        for cell in ws[1]:
            if isinstance(cell.value, str) and cell.value.strip().startswith("Delta"):
                cell.fill = fill


def _borders_and_stripes(ws: Worksheet) -> None:
    """Borda fina em tudo + zebra apenas nas linhas pares (depois do cabeçalho)."""
    thin   = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    max_r, max_c = ws.max_row, ws.max_column
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c), start=1):
        is_zebra = (r_idx > 1 and r_idx % 2 == 0)  # só pares e não cabeçalho
        for cell in row:
            cell.border = border
            if is_zebra:
                cell.fill = ZEBRA_FILL
            # sem else para não apagar fills existentes (ex.: cabeçalho)

def _space_items_xml(ws: Worksheet) -> None:
    """Insere 2 linhas em branco sempre que a CHAVE mudar (nova NF-e)."""
    if ws.max_row < 3:
        return
    header = [c.value for c in ws[1]]
    try:
        chave_col = header.index("CHAVE") + 1  # 1-based
    except ValueError:
        return
    r = ws.max_row
    while r > 2:
        atual = ws.cell(row=r,   column=chave_col).value
        acima = ws.cell(row=r-1, column=chave_col).value
        if atual and acima and atual != acima:
            ws.insert_rows(r, amount=2)
        r -= 1

# -------------------------
# Orquestração (única)
# -------------------------

def format_workbook(wb: Workbook) -> None:
    """Aplica formatação padrão em todas as abas + regras específicas."""
    for ws in wb.worksheets:
        _header_and_row_heights(ws)
        _auto_width(ws)
        _freeze_header(ws)
        _center_all(ws)
        _highlight_delta_headers(ws)
        _borders_and_stripes(ws)

    # Regra específica: aba “Itens (XML)” recebe espaçamento entre notas (por CHAVE)
    if "Itens (XML)" in wb.sheetnames:
        _space_items_xml(wb["Itens (XML)"])
