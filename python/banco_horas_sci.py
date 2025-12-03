"""
Banco de Horas SCI
Ficha de Horas Trabalhadas - Versão Completa
Usa verba 5 para horas trabalhadas
Usa verbas 603, 605, 608, 613, 615 para horas extras
"""

from core.connection import SCIConnection
import pandas as pd
from decimal import Decimal
from datetime import datetime, date
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def buscar_empresa_por_cnpj(cnpj: str, conn: SCIConnection):
    """Busca empresa pelo CNPJ"""
    cnpj_limpo = cnpj.replace('.', '').replace('/', '').replace('-', '')
    
    sql = f"""
    SELECT FIRST 1 BDCODEMP, BDNOMEMP, BDCNPJEMP 
    FROM VW_TEMPRESAS_REF 
    WHERE REPLACE(REPLACE(REPLACE(BDCNPJEMP, '.', ''), '/', ''), '-', '') = '{cnpj_limpo}'
    ORDER BY BDCODEMP
    """
    
    resultado = conn.execute_query(sql)
    if resultado:
        return {
            'codigo': resultado[0][0],
            'nome': resultado[0][1],
            'cnpj': resultado[0][2]
        }
    return None


def formatar_horas(valor) -> str:
    """Converte horas decimais para formato HH:MM"""
    if valor is None or valor == 0:
        return "0:00"
    
    # Converter Decimal para float se necessário
    if isinstance(valor, Decimal):
        valor = float(valor)
    elif not isinstance(valor, (int, float)):
        try:
            valor = float(valor)
        except:
            return "0:00"
    
    horas = int(valor)
    minutos = int((valor - horas) * 60)
    return f"{horas}:{minutos:02d}"


def formatar_horas_extras(valor) -> str:
    """Formata horas extras mantendo parte decimal como minutos diretos (ex: 1.14 = 1:14)
    Se minutos >= 60, converte para horas adicionais (ex: 10.81 = 11:21)"""
    if valor is None or valor == 0:
        return "0:00"
    
    # Converter Decimal para float se necessário
    if isinstance(valor, Decimal):
        valor = float(valor)
    elif not isinstance(valor, (int, float)):
        try:
            valor = float(valor)
        except:
            return "0:00"
    
    # Separar parte inteira (horas) e parte decimal (minutos diretos)
    horas = int(valor)
    parte_decimal = valor - horas  # Ex: 10.81 -> 0.81
    minutos = int(parte_decimal * 100)  # Ex: 0.81 * 100 = 81 minutos
    
    # Se minutos >= 60, converter para horas adicionais
    if minutos >= 60:
        horas_extra = minutos // 60  # Quantas horas completas nos minutos
        minutos_restantes = minutos % 60  # Minutos que sobram
        horas += horas_extra
        minutos = minutos_restantes
    
    return f"{horas}:{minutos:02d}"


def converter_horas_para_decimal(horas_str: str) -> float:
    """Converte formato HH:MM para horas decimais (ex: '220:30' -> 220.5)"""
    if not horas_str or horas_str == '0:00' or horas_str == '':
        return 0.0
    
    try:
        partes = horas_str.split(':')
        if len(partes) != 2:
            return 0.0
        horas = int(partes[0])
        minutos = int(partes[1])
        return horas + (minutos / 60.0)
    except:
        return 0.0


def converter_extras_para_decimal(extras_str: str) -> float:
    """Converte formato HH:MM de horas extras para horas decimais
    Considera que o formato original era decimal direto (ex: 1.14 = 1h14min)"""
    if not extras_str or extras_str == '0:00' or extras_str == '':
        return 0.0
    
    try:
        partes = extras_str.split(':')
        if len(partes) != 2:
            return 0.0
        horas = int(partes[0])
        minutos = int(partes[1])
        return horas + (minutos / 60.0)
    except:
        return 0.0


def data_para_competencia(data: date) -> int:
    """Converte data para formato de competência YYYYMM"""
    return data.year * 100 + data.month


def competencia_para_data(competencia: int) -> date:
    """Converte competência YYYYMM para data (primeiro dia do mês)"""
    ano = competencia // 100
    mes = competencia % 100
    return date(ano, mes, 1)


def obter_competencias_periodo(data_inicial: date, data_final: date) -> list:
    """Retorna lista de competências (YYYYMM) no período"""
    competencias = []
    data_atual = date(data_inicial.year, data_inicial.month, 1)
    data_fim = date(data_final.year, data_final.month, 1)
    
    while data_atual <= data_fim:
        competencias.append(data_para_competencia(data_atual))
        # Próximo mês
        if data_atual.month == 12:
            data_atual = date(data_atual.year + 1, 1, 1)
        else:
            data_atual = date(data_atual.year, data_atual.month + 1, 1)
    
    return competencias


def buscar_colaboradores_ativos_periodo(cod_emp: int, data_inicial: date, data_final: date, conn: SCIConnection):
    """Busca colaboradores que estavam ativos no período informado
    Considera ativo quem foi admitido antes ou durante o período e tem registros de folha no período
    (verba 5 ou horas extras - verbas 603, 605, 608, 613, 615)"""
    # Converter datas para formato do banco
    data_final_str = data_final.strftime('%Y-%m-%d')
    
    # Obter competências do período
    competencias = obter_competencias_periodo(data_inicial, data_final)
    competencia_inicial = competencias[0]
    competencia_final = competencias[-1]
    
    # OTIMIZAÇÃO: Buscar colaboradores que têm verba 5 OU horas extras no período
    # Usar UNION para combinar os dois resultados de forma eficiente
    # Isso é muito mais rápido que fazer JOIN sem filtro de verba
    verbas_extras = [603, 605, 608, 613, 615]
    verbas_str = ','.join(map(str, verbas_extras))
    
    sql = f"""
    SELECT DISTINCT 
        C.BDCODCOL,
        C.BDNOMCOL,
        C.BDDATAADMCOL
    FROM VW_COLABORADORES C
    INNER JOIN VWRH_RELATORIORECIBOFOLHA01 F
        ON C.BDCODEMP = F.BDCODEMP
        AND C.BDCODCOL = F.BDCODCOL
    WHERE C.BDCODEMP = {cod_emp}
    AND C.BDDATAADMCOL IS NOT NULL
    AND C.BDDATAADMCOL <= '{data_final_str}'
    AND (
        F.BDCODVER = 5
        OR F.BDCODVER IN ({verbas_str})
    )
    AND F.BDREFFN >= {competencia_inicial} AND F.BDREFFN <= {competencia_final}
    ORDER BY C.BDCODCOL
    """
    
    resultado = conn.execute_query(sql)
    colaboradores = []
    if resultado:
        for row in resultado:
            colaboradores.append({
                'codigo': row[0],
                'nome': row[1] or '',
                'data_admissao': row[2]
            })
    
    return colaboradores


def gerar_ficha_horas(cnpj: str, data_inicial: date = None, data_final: date = None, ano: int = None):
    """Gera a ficha completa de horas trabalhadas
    Pode ser chamada com período (data_inicial, data_final) ou com ano (para compatibilidade)"""
    conn = SCIConnection()
    
    # Se não foi informado período, usar ano (compatibilidade com código antigo)
    if data_inicial is None or data_final is None:
        if ano is None:
            print("[ERRO] Deve informar periodo (data_inicial e data_final) ou ano")
            return None
        data_inicial = date(ano, 1, 1)
        data_final = date(ano, 12, 31)
    
    tempo_inicio = time.time()
    
    print("=" * 80)
    print("GERACAO DE FICHA DE HORAS TRABALHADAS")
    print(f"CNPJ: {cnpj}")
    print(f"Periodo: {data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}")
    print("=" * 80)
    
    # Buscar empresa
    print("\n[LOG] 1. Buscando empresa...")
    tempo_etapa = time.time()
    empresa = buscar_empresa_por_cnpj(cnpj, conn)
    if not empresa:
        print("[ERRO] Empresa nao encontrada")
        return
    print(f"[OK] {empresa['nome']} (Codigo: {empresa['codigo']}) - Tempo: {time.time() - tempo_etapa:.2f}s")
    
    # Obter competências do período
    print("\n[LOG] Calculando competencias do periodo...")
    tempo_etapa = time.time()
    competencias = obter_competencias_periodo(data_inicial, data_final)
    competencia_inicial = competencias[0]
    competencia_final = competencias[-1]
    print(f"[INFO] Periodo convertido para competencias: {competencia_inicial} a {competencia_final}")
    print(f"[INFO] Total de meses: {len(competencias)} - Tempo: {time.time() - tempo_etapa:.2f}s")
    
    # Buscar colaboradores ativos no período
    print("\n[LOG] 2. Buscando colaboradores ativos no periodo...")
    tempo_etapa = time.time()
    colaboradores_ativos = buscar_colaboradores_ativos_periodo(
        empresa['codigo'], data_inicial, data_final, conn
    )
    
    if not colaboradores_ativos:
        print("[AVISO] Nenhum colaborador ativo encontrado no periodo")
        return None
    
    print(f"[OK] {len(colaboradores_ativos)} colaboradores ativos no periodo - Tempo: {time.time() - tempo_etapa:.2f}s")
    
    # Criar lista de códigos de colaboradores para filtrar
    print("\n[LOG] Preparando lista de colaboradores para consulta...")
    tempo_etapa = time.time()
    codigos_colaboradores = [col['codigo'] for col in colaboradores_ativos]
    codigos_str = ','.join(map(str, codigos_colaboradores))
    print(f"[INFO] {len(codigos_colaboradores)} colaboradores na lista - Tempo: {time.time() - tempo_etapa:.2f}s")
    
    # Buscar horas trabalhadas (verba 5) - apenas para colaboradores ativos no período
    print("\n[LOG] 3. Buscando horas trabalhadas (verba 5)...")
    print(f"[INFO] Executando query SQL para verba 5...")
    tempo_etapa = time.time()
    
    sql_horas = f"""
    SELECT 
        BDCODCOL,
        BDNOMCOL,
        BDREFFN,
        BDREFVFN_SUM
    FROM VWRH_RELATORIORECIBOFOLHA01
    WHERE BDCODEMP = {empresa['codigo']}
    AND BDCODVER = 5
    AND BDREFFN >= {competencia_inicial} AND BDREFFN <= {competencia_final}
    AND BDCODCOL IN ({codigos_str})
    ORDER BY BDCODCOL, BDREFFN
    """
    
    resultado_horas = conn.execute_query(sql_horas)
    tempo_query = time.time() - tempo_etapa
    
    if not resultado_horas:
        print("[AVISO] Nenhuma hora trabalhada encontrada")
        resultado_horas = []
    else:
        print(f"[OK] {len(resultado_horas)} registros de horas trabalhadas - Tempo da query: {tempo_query:.2f}s")
    
    # Buscar horas extras (verbas 603, 605, 608, 613, 615) - apenas para colaboradores ativos
    print("\n[LOG] 4. Buscando horas extras (verbas 603, 605, 608, 613, 615)...")
    print(f"[INFO] Executando query SQL para horas extras...")
    tempo_etapa = time.time()
    
    verbas_extras = [603, 605, 608, 613, 615]
    verbas_str = ','.join(map(str, verbas_extras))
    
    sql_extras = f"""
    SELECT 
        BDCODCOL,
        BDNOMCOL,
        BDREFFN,
        BDREFVFN_SUM,
        BDCODVER
    FROM VWRH_RELATORIORECIBOFOLHA01
    WHERE BDCODEMP = {empresa['codigo']}
    AND BDCODVER IN ({verbas_str})
    AND BDREFFN >= {competencia_inicial} AND BDREFFN <= {competencia_final}
    AND BDCODCOL IN ({codigos_str})
    ORDER BY BDCODCOL, BDREFFN, BDCODVER
    """
    
    resultado_extras = conn.execute_query(sql_extras)
    tempo_query = time.time() - tempo_etapa
    
    if not resultado_extras:
        print("[AVISO] Nenhuma hora extra encontrada")
        resultado_extras = []
    else:
        print(f"[OK] {len(resultado_extras)} registros de horas extras - Tempo da query: {tempo_query:.2f}s")
    
    # Organizar dados por colaborador e mês
    print("\n[LOG] 5. Processando dados dos colaboradores...")
    tempo_etapa = time.time()
    
    dados_colaboradores = {}
    meses_nomes = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
    
    # Inicializar todos os colaboradores ativos
    print(f"[INFO] Inicializando estrutura de dados para {len(colaboradores_ativos)} colaboradores...")
    tempo_sub = time.time()
    for colab in colaboradores_ativos:
        dados_colaboradores[colab['codigo']] = {
            'nome': colab['nome'],
            'horas_meses': {},
            'extras_meses': {}
        }
    print(f"[INFO] Estrutura inicializada - Tempo: {time.time() - tempo_sub:.2f}s")
    
    # Processar horas trabalhadas - converter dias para horas
    print(f"[INFO] Processando {len(resultado_horas)} registros de horas trabalhadas...")
    tempo_sub = time.time()
    contador = 0
    for row in resultado_horas:
        cod_col = row[0]
        competencia = row[2]
        ref_horas = row[3] if row[3] else 0
        
        mes = competencia % 100
        
        # Converter referência (dias) para horas: 30 dias = 220 horas
        if isinstance(ref_horas, Decimal):
            dias = float(ref_horas)
        else:
            dias = float(ref_horas) if ref_horas else 0.0
        
        # Se parece ser dias (1-31), converter para horas
        if 1 <= dias <= 31:
            horas = dias * (220.0 / 30.0)  # 30 dias = 220 horas
        else:
            horas = dias
        
        dados_colaboradores[cod_col]['horas_meses'][mes] = horas
        contador += 1
        if contador % 100 == 0:
            print(f"[INFO] Processados {contador}/{len(resultado_horas)} registros de horas trabalhadas...")
    
    print(f"[INFO] Horas trabalhadas processadas - Tempo: {time.time() - tempo_sub:.2f}s")
    
    # Processar horas extras - usar campo REFERENCIA diretamente (não somar durante processamento)
    print(f"[INFO] Processando {len(resultado_extras)} registros de horas extras...")
    tempo_sub = time.time()
    contador = 0
    for row in resultado_extras:
        cod_col = row[0]
        competencia = row[2]
        ref_extras = row[3] if row[3] else 0
        
        mes = competencia % 100
        
        # Usar referência diretamente (sem conversão)
        if isinstance(ref_extras, Decimal):
            horas_extras_float = float(ref_extras)
        else:
            horas_extras_float = float(ref_extras) if ref_extras else 0.0
        
        # Armazenar cada valor de hora extra separadamente (não somar)
        # Vamos somar apenas na exibição final
        if mes not in dados_colaboradores[cod_col]['extras_meses']:
            dados_colaboradores[cod_col]['extras_meses'][mes] = []
        
        dados_colaboradores[cod_col]['extras_meses'][mes].append(horas_extras_float)
        contador += 1
        if contador % 100 == 0:
            print(f"[INFO] Processados {contador}/{len(resultado_extras)} registros de horas extras...")
    
    print(f"[INFO] Horas extras processadas - Tempo: {time.time() - tempo_sub:.2f}s")
    print(f"[OK] {len(dados_colaboradores)} colaboradores processados - Tempo total: {time.time() - tempo_etapa:.2f}s")
    
    # Criar estrutura de dados para DataFrame
    print("\n[LOG] 6. Montando tabela para exportacao...")
    tempo_etapa = time.time()
    dados = []
    
    # Obter meses do período para criar colunas dinâmicas
    print("[INFO] Preparando colunas do periodo...")
    meses_periodo = []
    for comp in competencias:
        mes_num = comp % 100
        meses_periodo.append((mes_num, meses_nomes[mes_num - 1]))
    print(f"[INFO] {len(meses_periodo)} meses no periodo")
    
    print(f"[INFO] Processando {len(dados_colaboradores)} colaboradores para gerar linhas da tabela...")
    contador_linhas = 0
    for cod_col, info in sorted(dados_colaboradores.items()):
        linha = {
            'cnpj_empresa': empresa['cnpj'],
            'razao_social_empresa': empresa['nome'],
            'codigo_centro_custo': '',
            'descricao_centro_custo': '',
            'matricula_colaborador': cod_col,
            'nome_colaborador': info['nome'],
            'carga_horaria_regime': '220:00'
        }
        
        total_horas = 0.0
        total_extras_decimal = 0.0  # Soma já convertida para horas decimais
        
        # Função auxiliar para converter horas extras para horas decimais
        def converter_extras_para_horas_decimais(valor_extras):
            """Converte horas extras (formato 1.14 = 1h14min) para horas decimais"""
            if valor_extras == 0:
                return 0.0
            horas = int(valor_extras)
            parte_decimal = valor_extras - horas
            minutos = int(parte_decimal * 100)
            # Se minutos >= 60, já foi tratado na formatação, mas aqui precisamos considerar
            if minutos >= 60:
                horas_extra = minutos // 60
                minutos_restantes = minutos % 60
                horas += horas_extra
                minutos = minutos_restantes
            # Converter minutos para fração de hora
            return horas + (minutos / 60.0)
        
        # Adicionar horas apenas para os meses do período
        for mes_num, mes_nome in meses_periodo:
            horas = info['horas_meses'].get(mes_num, 0.0)
            if isinstance(horas, Decimal):
                horas = float(horas)
            
            # Para horas extras, somar apenas na exibição
            extras_lista = info['extras_meses'].get(mes_num, [])
            extras_soma = sum(extras_lista) if extras_lista else 0.0
            
            linha[f'horas_trabalhadas_{mes_nome}'] = formatar_horas(horas)
            linha[f'horas_extras_{mes_nome}'] = formatar_horas_extras(extras_soma)
            
            # Somar horas trabalhadas
            total_horas += horas
            
            # Converter horas extras para decimal e somar
            extras_decimal = converter_extras_para_horas_decimais(extras_soma)
            total_extras_decimal += extras_decimal
        
        linha['total_horas_trabalhadas'] = formatar_horas(total_horas)
        
        # Para exibir total de horas extras, precisamos converter de volta
        # Mas como não podemos converter facilmente de horas decimais para o formato especial,
        # vamos somar os valores decimais originais e formatar
        total_extras_bruto = 0.0
        for mes_num, _ in meses_periodo:
            extras_lista = info['extras_meses'].get(mes_num, [])
            extras_soma = sum(extras_lista) if extras_lista else 0.0
            total_extras_bruto += extras_soma
        
        linha['total_horas_extras'] = formatar_horas_extras(total_extras_bruto)
        
        # Calcular total_geral usando a soma já convertida para horas decimais
        total_geral_decimal = total_horas + total_extras_decimal
        linha['total_geral'] = formatar_horas(total_geral_decimal)
        
        dados.append(linha)
        contador_linhas += 1
        if contador_linhas % 20 == 0:
            print(f"[INFO] Processadas {contador_linhas}/{len(dados_colaboradores)} linhas...")
    
    print(f"[INFO] Todas as linhas processadas - Tempo: {time.time() - tempo_etapa:.2f}s")
    
    # Criar DataFrame
    print("[INFO] Criando DataFrame do pandas...")
    tempo_sub = time.time()
    df = pd.DataFrame(dados)
    print(f"[INFO] DataFrame criado com {len(df)} linhas e {len(df.columns)} colunas - Tempo: {time.time() - tempo_sub:.2f}s")
    
    # Reordenar colunas - estrutura: Horas Trabalhadas Jan | Horas Extras Jan | Horas Trabalhadas Fev | Horas Extras Fev | ...
    colunas_base = [
        'cnpj_empresa', 'razao_social_empresa', 'codigo_centro_custo', 'descricao_centro_custo',
        'matricula_colaborador', 'nome_colaborador', 'carga_horaria_regime'
    ]
    
    colunas_meses = []
    for mes_num, mes_nome in meses_periodo:
        colunas_meses.extend([f'horas_trabalhadas_{mes_nome}', f'horas_extras_{mes_nome}'])
    
    colunas_finais = colunas_base + colunas_meses + ['total_horas_trabalhadas', 'total_horas_extras', 'total_geral']
    
    # Garantir que todas as colunas existem
    for col in colunas_finais:
        if col not in df.columns:
            df[col] = ""
    
    df = df[colunas_finais]
    
    # Exportar para Excel (planilha completa)
    print("\n[LOG] 7. Exportando planilha completa para Excel...")
    tempo_etapa = time.time()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    periodo_str = f"{data_inicial.strftime('%Y%m%d')}_{data_final.strftime('%Y%m%d')}"
    nome_arquivo_completo = f"Banco_Horas_SCI_{empresa['codigo']}_{periodo_str}_{timestamp}.xlsx"
    print(f"[INFO] Gerando arquivo completo: {nome_arquivo_completo}")
    df.to_excel(nome_arquivo_completo, index=False, engine='openpyxl')
    print(f"[OK] Arquivo completo gerado - Tempo: {time.time() - tempo_etapa:.2f}s")
    
    # Criar planilha enxuta e formatada
    nome_arquivo_formatado = criar_planilha_enxuta_formatada(nome_arquivo_completo, meses_periodo)
    
    tempo_total = time.time() - tempo_inicio
    print(f"\n{'='*80}")
    print(f"RESUMO DA GERACAO")
    print(f"{'='*80}")
    print(f"Total de registros: {len(df)}")
    print(f"Total de colunas (completo): {len(df.columns)}")
    print(f"Arquivo completo: {nome_arquivo_completo}")
    print(f"Arquivo formatado: {nome_arquivo_formatado}")
    print(f"Tempo total de execucao: {tempo_total:.2f}s ({tempo_total/60:.2f} minutos)")
    print(f"{'='*80}\n")
    
    # Retornar DataFrame (mantém compatibilidade com código existente)
    return df


def criar_planilha_enxuta_formatada(nome_arquivo_completo: str, meses_periodo: list) -> str:
    """
    Cria uma planilha enxuta e formatada a partir da planilha completa.
    Consolida Horas Trabalhadas + Horas Extras em uma única coluna por mês.
    
    Args:
        nome_arquivo_completo: Caminho da planilha completa
        meses_periodo: Lista de tuplas (mes_num, mes_nome) do período
    
    Returns:
        Caminho da nova planilha formatada
    """
    print("\n[LOG] 8. Criando planilha enxuta e formatada...")
    tempo_etapa = time.time()
    
    # Ler planilha completa
    df_completo = pd.read_excel(nome_arquivo_completo, engine='openpyxl')
    
    # Criar novo DataFrame enxuto
    dados_enxutos = []
    
    # Colunas base (mesmas da planilha completa)
    colunas_base = [
        'cnpj_empresa', 'razao_social_empresa', 'codigo_centro_custo', 
        'descricao_centro_custo', 'matricula_colaborador', 'nome_colaborador', 
        'carga_horaria_regime'
    ]
    
    for _, row in df_completo.iterrows():
        linha_enxuta = {}
        
        # Copiar colunas base
        for col in colunas_base:
            linha_enxuta[col] = row.get(col, '')
        
        # Consolidar horas por mês (Horas Trabalhadas + Horas Extras)
        total_horas_trabalhadas_decimal = 0.0  # Apenas horas trabalhadas (sem extras)
        total_geral_decimal = 0.0  # Horas trabalhadas + extras
        
        for mes_num, mes_nome in meses_periodo:
            col_horas = f'horas_trabalhadas_{mes_nome}'
            col_extras = f'horas_extras_{mes_nome}'
            
            # Obter valores formatados
            horas_str = str(row.get(col_horas, '0:00'))
            extras_str = str(row.get(col_extras, '0:00'))
            
            # Converter para decimal
            horas_decimal = converter_horas_para_decimal(horas_str)
            extras_decimal = converter_extras_para_decimal(extras_str)
            
            # Somar para total geral
            total_mes_decimal = horas_decimal + extras_decimal
            total_geral_decimal += total_mes_decimal
            total_horas_trabalhadas_decimal += horas_decimal  # Apenas horas trabalhadas
            
            # Formatar e adicionar à linha (apenas uma coluna por mês)
            linha_enxuta[mes_nome.upper()] = formatar_horas(total_mes_decimal)
        
        # Adicionar total de horas trabalhadas (sem extras)
        linha_enxuta['TOTAL_HORAS_TRABALHADAS'] = formatar_horas(total_horas_trabalhadas_decimal)
        
        # Adicionar total geral (com extras)
        linha_enxuta['TOTAL'] = formatar_horas(total_geral_decimal)
        
        dados_enxutos.append(linha_enxuta)
    
    # Criar DataFrame enxuto
    df_enxuto = pd.DataFrame(dados_enxutos)
    
    # Reordenar colunas: base + meses + total_horas_trabalhadas + total
    colunas_meses = [mes_nome.upper() for _, mes_nome in meses_periodo]
    colunas_ordenadas = colunas_base + colunas_meses + ['TOTAL_HORAS_TRABALHADAS', 'TOTAL']
    
    # Garantir que todas as colunas existem
    for col in colunas_ordenadas:
        if col not in df_enxuto.columns:
            df_enxuto[col] = ""
    
    df_enxuto = df_enxuto[colunas_ordenadas]
    
    # Nome do arquivo formatado
    nome_arquivo_formatado = nome_arquivo_completo.replace('.xlsx', '_FORMATADO.xlsx')
    
    # Exportar para Excel
    print(f"[INFO] Gerando planilha formatada: {nome_arquivo_formatado}")
    df_enxuto.to_excel(nome_arquivo_formatado, index=False, engine='openpyxl')
    
    # Aplicar formatação
    formatar_excel(nome_arquivo_formatado, meses_periodo)
    
    print(f"[OK] Planilha formatada gerada - Tempo: {time.time() - tempo_etapa:.2f}s")
    
    return nome_arquivo_formatado


def formatar_excel(nome_arquivo: str, meses_periodo: list):
    """Formata a planilha Excel com estilos profissionais"""
    print("[INFO] Aplicando formatação à planilha...")
    
    # Carregar workbook
    wb = load_workbook(nome_arquivo)
    ws = wb.active
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Side(style='thin', color='000000')
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Formatar cabeçalho (linha 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Ajustar largura das colunas
    larguras_padrao = {
        'A': 18,  # CNPJ
        'B': 40,  # Razão Social
        'C': 15,  # Código Centro Custo
        'D': 30,  # Descrição Centro Custo
        'E': 12,  # Matrícula
        'F': 30,  # Nome
        'G': 15,  # Carga Horária
    }
    
    # Aplicar larguras padrão
    for col_letter, width in larguras_padrao.items():
        ws.column_dimensions[col_letter].width = width
    
    # Largura para colunas de meses (mais estreitas)
    col_inicial_meses = 8  # Coluna H (após carga_horaria_regime)
    for idx, (_, mes_nome) in enumerate(meses_periodo):
        col_letter = get_column_letter(col_inicial_meses + idx)
        ws.column_dimensions[col_letter].width = 12
    
    # Largura para colunas de totais
    col_total_horas = get_column_letter(col_inicial_meses + len(meses_periodo))
    col_total_geral = get_column_letter(col_inicial_meses + len(meses_periodo) + 1)
    ws.column_dimensions[col_total_horas].width = 20  # Total Horas Trabalhadas
    ws.column_dimensions[col_total_geral].width = 15  # Total Geral
    
    # Congelar primeira linha e primeiras colunas (até nome)
    ws.freeze_panes = 'H2'  # Congela até coluna G (nome) e linha 1
    
    # Aplicar bordas e alinhamento nas células de dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            
            # Identificar tipo de coluna
            col_name = ws[f'{get_column_letter(cell.column)}1'].value
            
            if col_name and isinstance(col_name, str):
                col_name_upper = col_name.upper()
                # Colunas de horas (meses e totais) - alinhar à direita
                if (col_name_upper in [mes.upper() for _, mes in meses_periodo] or 
                    col_name_upper in ['TOTAL_HORAS_TRABALHADAS', 'TOTAL']):
                    cell.alignment = right_align
                # Colunas numéricas (matrícula, código centro custo) - alinhar à direita
                elif col_name_upper in ['MATRICULA_COLABORADOR', 'CODIGO_CENTRO_CUSTO']:
                    cell.alignment = right_align
                # Demais colunas - alinhar à esquerda
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Altura da linha do cabeçalho
    ws.row_dimensions[1].height = 25
    
    # Salvar
    wb.save(nome_arquivo)
    print("[OK] Formatação aplicada com sucesso")


